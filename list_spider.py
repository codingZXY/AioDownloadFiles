import requests
import threading
from openpyxl import Workbook
from parsel import Selector
from setting import *
from collections import OrderedDict
import traceback
import time
import os
import json
import re
from urllib.parse import unquote

HEADER_COLUMNS = OrderedDict({
    'title':'标题',
    'category':'分类',
    'publisher':'发布者',
    'start_date':'起始日期',
    'end_date':'截止日期',
})

class ListSpider():
    def __init__(self, tasks):
        self.tasks = tasks
        self.headers = {
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,und;q=0.7',
            'Connection': 'keep-alive',
            'Host': 'www.c-whale.com',
            'Referer': 'http://www.c-whale.com/index',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36'
        }
        self.failed = []

    def crawl(self, url):
        try:
            resp = requests.get(url, headers=self.headers)
            if resp.status_code == 200:
                return resp.text
            else:
                self.failed.append({
                    'code': resp.status_code,
                    'msg': 'wrong status',
                    'url': url,
                })
        except:
            self.failed.append({
                'code': '',
                'msg': traceback.format_exc(),
                'url': url,
            })

    def clean(self, source):
        items = []
        sel = Selector(source)
        data_list = sel.xpath('//div[@class="newlist"]')
        for data in data_list:
            title = str(data.xpath('string(./a[@class="conn"]/@href)').re_first('tit=(.*)')).strip()
            title = unquote(unquote(title))
            category = str(data.xpath('string(.//a[@class="p-level lable"])').get()).strip()
            publisher = str(data.xpath('string(.//a[@class="nature lable"])').get()).strip()
            start_date = str(data.xpath('string(.//span[@class="time"][1])').get()).strip().lstrip('起始日期：')
            end_date = str(data.xpath('string(.//span[@class="time"][2])').get()).strip().lstrip('截止日期：')

            items.append({
                'title':title,
                'category':category,
                'publisher':publisher,
                'start_date':start_date,
                'end_date':end_date,
            })

        return items

    def save_rows(self, rows_data, ws):
        for row_data in rows_data:
            row = [row_data[col] for col in HEADER_COLUMNS.keys()]
            ws.append(row)

    def set_file_link(self, ws, dir):
        ws.cell(row = 1, column = 6).value = '文件位置'
        for r in range(2, ws.max_row+1):
            title = ws.cell(row=r,column=1).value
            file_name = re.sub(r'[\\/:*?"<>|]', '-', f'{title}.pdf')
            file_name = f'{dir}/{file_name}'
            if os.path.exists(file_name):
                ws.cell(row=r, column=6).value = '= HYPERLINK("{}","{}")'.format(file_name, file_name)

    def save(self, wb, filename):
        wb.save(f"{filename}.xlsx")

    def run(self, task):
        wb = Workbook()
        ws = wb.active
        ws.append([col for col in HEADER_COLUMNS.values()])

        aid = TASK_TYPE_ID[task]
        filename = TASK_TYPE_NAME[task]
        page = 1
        while 1:
            print(f'正在采集:{filename} 第 {page} 页')
            url = f'http://www.c-whale.com/jsp/list?aid={aid}&bid=&row={page*10}'
            source = self.crawl(url)
            if source is None:
                print(f'出错:{filename} 第 {page} 页, 已跳过')
                continue

            rows_data = self.clean(source)
            if not rows_data:
                print(f'{filename} 第 {page} 页无数据')
                break

            self.save_rows(rows_data, ws)

            page += 1
            time.sleep(1)

        self.set_file_link(ws,filename)
        self.save(wb, filename)

    def save_json(self):
        dir_name = 'log'
        if not os.path.exists(dir_name):
            os.mkdir(dir_name)

        file_name =f'{dir_name}/{int(time.time()*1000)}_list.json'
        with open(file_name,'w') as f:
            json.dump(self.failed,f)

    def start(self):
        print('开始'.center(100, '-'))
        start = int(time.time())

        threads = []
        for task in self.tasks:
            print(f'开始采集:{TASK_TYPE_NAME[task]}')
            t = threading.Thread(target=self.run, args=(task,), name=task)
            t.setDaemon(True)
            t.start()
            threads.append(t)

        for t in threads:
            t.join()
            print(f'结束采集:{TASK_TYPE_NAME[t.getName()]}')

        end = int(time.time())
        total_time = int((end - start))

        print(f'耗时:{total_time} 秒')
        print('结束'.center(100, '-'))

        self.save_json()


if __name__ == '__main__':
    # tasks = ['0']
    # spider = ListSpider(tasks)
    # source = spider.crawl('http://www.c-whale.com/jsp/list?aid=%C3%A7%C2%94%C2%B3%C3%A6%C2%8A%C2%A5%C3%A9%C2%80%C2%9A%C3%A7%C2%9F%C2%A5&bid=&row=10')
    # spider.clean(source)
    tasks = ['1']
    spider = ListSpider(tasks)
    spider.start()

